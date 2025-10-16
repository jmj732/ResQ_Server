from sqlmodel import Session, select, func
from app.question.model import InterviewQuestion, Company
from app.question.dto.question_dto import (
    QuestionResponse,
    QuestionAdditionResponse,
    QuestionUpdateResponse,
    QuestionDeleteResponse
)
from app.common.dto.page_dto import Page, Pageable
from typing import Optional
import math


class QuestionService:
    def __init__(self, session: Session):
        self.session = session

    def get_companies(self) -> list[Company]:
        """Get all companies with their IDs and names"""
        statement = select(Company.id, Company.name)
        results = self.session.exec(statement).all()
        return [Company(id=company_id, name=name) for company_id, name in results]

    def search_questions(
        self,
        company_id: Optional[int] = None,
        year: Optional[int] = None,
        category: Optional[str] = None,
        pageable: Pageable = Pageable()
    ) -> Page[QuestionResponse]:
        """Search interview questions with filters and pagination (like Spring Page)"""

        # Base query
        statement = select(InterviewQuestion, Company).join(
            Company, Company.id == InterviewQuestion.company_id
        )

        # Apply filters
        if company_id is not None and company_id > 0:
            statement = statement.where(InterviewQuestion.company_id == company_id)

        if year is not None:
            statement = statement.where(InterviewQuestion.year == year)

        if category is not None:
            statement = statement.where(InterviewQuestion.category == category)

        # Count total elements (before pagination)
        count_statement = select(func.count()).select_from(InterviewQuestion)
        if company_id is not None:
            count_statement = count_statement.where(InterviewQuestion.company_id == company_id)
        if year is not None:
            count_statement = count_statement.where(InterviewQuestion.year == year)
        if category is not None:
            count_statement = count_statement.where(InterviewQuestion.category == category)

        total_elements = self.session.exec(count_statement).one()

        # Apply pagination
        statement = statement.offset(pageable.get_offset()).limit(pageable.get_limit())

        # Execute query
        results = self.session.exec(statement).all()

        # Transform to response
        content = []
        for question, company in results:
            content.append(QuestionResponse(
                id=question.id,
                company_id=question.company_id,
                company_name=company.name,
                year=question.year,
                category=question.category,
                question=question.question
            ))

        # Calculate page info
        total_pages = math.ceil(total_elements / pageable.size) if total_elements > 0 else 0
        number_of_elements = len(content)

        return Page[QuestionResponse](
            content=content,
            pageable=pageable,
            size=pageable.size,
            number=pageable.page,
            number_of_elements=number_of_elements,
            first=(pageable.page == 0),
            last=(pageable.page >= total_pages - 1),
            empty=(number_of_elements == 0),
            total_elements=total_elements,
            total_pages=total_pages
        )

    def add_question(self, company_id: int, year: int,
                     category: Optional[str], question: Optional[str]) -> QuestionAdditionResponse:
        """Add a new interview question"""

        # Verify company exists
        company = self.session.get(Company, company_id)
        if not company:
            raise ValueError("Company not found")

        new_question = InterviewQuestion(
            company_id=company_id,
            year=year,
            category=category,
            question=question
        )

        self.session.add(new_question)
        self.session.flush()  # ID를 생성하지만 commit은 나중에

        return QuestionAdditionResponse(
            question_id=new_question.id
        )

    def update_question(self, question_id: int, content: str) -> QuestionUpdateResponse:
        """Update an interview question's content"""

        # Find the question
        question = self.session.get(InterviewQuestion, question_id)
        if not question:
            raise ValueError("Question not found")

        # Update the content
        question.question = content
        self.session.flush()

        return QuestionUpdateResponse(
            message="Question updated successfully"
        )

    def delete_question(self, question_id: int) -> QuestionDeleteResponse:
        """Delete an interview question"""

        # Find the question
        question = self.session.get(InterviewQuestion, question_id)
        if not question:
            raise ValueError("Question not found")

        # Delete the question
        self.session.delete(question)
        self.session.flush()

        return QuestionDeleteResponse(
            message="Question deleted successfully"
        )


    def create_sample_excel(self) -> bytes:
        """Create a sample Excel file for bulk upload"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "질문 업로드 양식"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ["회사명", "년도", "카테고리", "질문"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Sample data
        sample_data = [
            ["핀다", 2024, "front", "자기소개를 해주세요."],
            ["더스팟", 2024, "back", "우리 회사에 지원한 이유는 무엇인가요?"],
            ["더스팟", 2023, "AI", "최근에 진행한 프로젝트에 대해 설명해주세요."],
        ]
        
        for row_num, data in enumerate(sample_data, 2):
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.getvalue()

    def bulk_upload_questions(self, file_content: bytes, filename: str) -> dict:
        """Process bulk question upload from Excel file"""
        import io
        from openpyxl import load_workbook
        
        results = {
            "total_rows": 0,
            "success_count": 0,
            "error_count": 0,
            "errors": []
        }
        
        try:
            # Load Excel file
            wb = load_workbook(io.BytesIO(file_content))
            ws = wb.active
            
            # Skip header row
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            results["total_rows"] = len(rows)
            
            for idx, row in enumerate(rows, start=2):
                try:
                    # Validate row has data
                    if not any(row):
                        continue
                    
                    company_name, year, category, question = row[0], row[1], row[2], row[3]
                    
                    # Validate required fields
                    if not company_name or not year:
                        results["errors"].append({
                            "row": idx,
                            "error": "회사명과 년도는 필수 항목입니다."
                        })
                        results["error_count"] += 1
                        continue
                    
                    # Find company by name
                    company_statement = select(Company).where(Company.name == str(company_name).strip())
                    company = self.session.exec(company_statement).first()
                    
                    if not company:
                        results["errors"].append({
                            "row": idx,
                            "error": f"회사명 '{company_name}'을(를) 찾을 수 없습니다."
                        })
                        results["error_count"] += 1
                        continue
                    
                    # Create question
                    new_question = InterviewQuestion(
                        company_id=company.id,
                        year=int(year),
                        category=category if category else None,
                        question=question if question else None
                    )
                    
                    self.session.add(new_question)
                    results["success_count"] += 1
                    
                except Exception as e:
                    results["errors"].append({
                        "row": idx,
                        "error": f"처리 중 오류: {str(e)}"
                    })
                    results["error_count"] += 1
            
            # Commit all successful inserts
            if results["success_count"] > 0:
                self.session.commit()
            
        except Exception as e:
            results["errors"].append({
                "row": 0,
                "error": f"파일 처리 중 오류: {str(e)}"
            })
            results["error_count"] += 1
        
        return results

